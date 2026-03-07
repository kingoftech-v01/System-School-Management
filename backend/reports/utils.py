"""
Core PDF/CSV/Excel generation utilities for the reports app.
Extends the existing render_to_pdf pattern from accounts/views_frontend.py.
"""
import csv
import io
from datetime import datetime

from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa


def get_school_context(request):
    """Extract school branding data for PDF headers."""
    school = getattr(request, 'tenant', None)
    return {
        'school_name': getattr(school, 'name', 'School Management System'),
        'school_address': getattr(school, 'address', ''),
        'school_phone': getattr(school, 'phone', ''),
        'school_email': getattr(school, 'email', ''),
        'school_logo_url': (
            school.logo.url if school and hasattr(school, 'logo') and school.logo
            else None
        ),
    }


def render_report_pdf(
    template_name, context, request=None, filename='report.pdf',
    confidential=False, orientation='portrait',
):
    """
    Enhanced render_to_pdf with school branding, page numbers, metadata.
    """
    context.update({
        'generated_at': datetime.now(),
        'generated_by': (
            request.user.get_full_name
            if request and request.user.is_authenticated
            else 'System'
        ),
        'is_confidential': confidential,
        'page_orientation': orientation,
    })

    if request:
        context.update(get_school_context(request))

    html_string = render_to_string(template_name, context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    result = pisa.CreatePDF(
        io.BytesIO(html_string.encode('utf-8')),
        dest=response,
        encoding='utf-8',
    )

    if result.err:
        return HttpResponse('Error generating PDF report.', status=500)

    return response


def render_csv_response(filename, header_row, data_rows):
    """Generate a CSV response from data."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(header_row)
    for row in data_rows:
        writer.writerow(row)

    return response


def render_excel_response(filename, sheet_name, header_row, data_rows):
    """Generate an Excel response using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return HttpResponse(
            'Excel export requires openpyxl. Install it with: '
            'pip install openpyxl',
            status=500,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(
        start_color='0D6EFD', end_color='0D6EFD', fill_type='solid'
    )

    ws.append(header_row)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in data_rows:
        ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(
            max_length + 2, 50
        )

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet'
        )
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def log_export(
    request, report_type, export_format, title,
    student_id=None, incident_id=None, filter_params=None, reason='',
):
    """Record an export in the audit trail."""
    from .models import ReportExportLog

    ReportExportLog.objects.create(
        report_type=report_type,
        export_format=export_format,
        title=title,
        exported_by=(
            request.user if request.user.is_authenticated else None
        ),
        export_reason=reason,
        student_id=student_id,
        incident_id=incident_id,
        filter_params=filter_params or {},
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', ''),
        tenant=getattr(request, 'tenant', None),
    )
